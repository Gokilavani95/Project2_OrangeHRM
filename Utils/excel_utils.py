from openpyxl import load_workbook


class ExcelUtils:

    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name

        self.workbook = load_workbook(self.file_path)
        self.sheet = self.workbook[self.sheet_name]

    def get_login_data(self):
        """
        Reads all login data from Excel.
        Assumes:
        Username = column F
        Password = column G
        """

        test_data = []

        for row in range(2, self.sheet.max_row + 1):

            data = {
                "row": row,
                "sl_no": self.sheet.cell(row, 1).value,
                "test_id": self.sheet.cell(row, 2).value,
                "tester": self.sheet.cell(row, 3).value,
                "date": self.sheet.cell(row, 4).value,
                "test_parameter": self.sheet.cell(row, 5).value,
                "username": self.sheet.cell(row, 6).value,
                "password": self.sheet.cell(row, 7).value
            }

            test_data.append(data)

        return test_data

    def write_result(self, row_number, result, actual_result):
        """
        Writes Test Result and Actual Result back to Excel.
        """

        # Column H = Test Result
        self.sheet.cell(row_number, 8).value = result

        # Column I = Actual Result
        self.sheet.cell(row_number, 9).value = actual_result

        self.workbook.save(self.file_path)

    def close(self):
        self.workbook.close()